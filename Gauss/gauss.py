import openpyxl
import math
import matplotlib.pyplot as plt
from scipy.stats import norm
import statistics
import numpy as np

class Gauss:
    data = [] # tablica na dane
    n = 0 # ilość zmiennych
    avg = 0 # średnia wartość danych
    dev = 0 # odchylenie standardowe
    points = [] # zrzutowane punkty

    def __init__(self, minimum, maximum, step, path):
        self.minimum = minimum
        self.maximum = maximum
        self.step = step
        self.path = path


    #
    # pobieranie danych - OK
    #
    def get_data(self):
        # odczytanie danych
        dataframe_raw = openpyxl.load_workbook(f"datas/{self.path}")
        dataframe = dataframe_raw.active

        # liczba zmiennych
        self.n = dataframe.max_row-1

        # przypisanie danych do tablicy
        for row in range(1, dataframe.max_row):
            for col in dataframe.iter_cols(1, dataframe.max_column):
                self.data.append(col[row].value)

    #
    # średnia - OK
    #
    def average(self):
        # liczenie sredniej
        sum = 0
        for x in range(0, self.n):
            sum += self.data[x]
        self.avg = sum / self.n

    #
    # odchylenie standardowe - OK
    #
    def deviation(self):
        # odchylenie standardowe
        sum = 0
        for i in range(0, self.n):
            sum += pow((self.data[i] - self.avg), 2)
        self.dev = math.sqrt(sum / self.n)

    #
    # wzór Gaussa - OK
    #
    def gauss(self, x):
        return (1 / (self.dev * math.sqrt(2*math.pi))) * math.exp(((-1)*pow((x - self.avg), 2))/(2 * pow(self.dev, 2)))
    def point_projection(self):
        suma = 0
        for i in range(self.minimum, self.maximum+1):
            suma += self.data.count(i)
            # print(self.data.count(i))
            self.points.append(self.data.count(i)/self.n)
        # print(suma)
        # print(np.sum(self.points))

    #
    # rzutowanie całego wykresu - OK
    #
    def plot_write(self):
        # rysowanie krzywej gaussa
        x_axis = np.arange(self.minimum, self.maximum, 0.01)
        plt.plot(x_axis, norm.pdf(x_axis, self.avg, self.dev))

        # rzutowanie punktów na wykres
        # for i in range(0, len(self.points), self.step):
        #     plt.scatter(self.minimum+i, self.points[i], color="red")

        plt.savefig("plot.png")

        plt.show()

    #
    # metoda run - OK
    #
    def run(self):
        self.get_data()
        self.average()
        self.deviation()
        #self.point_projection()
        self.plot_write()
    def __str__(self) -> str:
        return f"min: {self.minimum}, max: {self.maximum}, path: {self.path}, n: {self.n}, avg: {self.avg}, dev: {self.dev}"
