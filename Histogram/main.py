import openpyxl
import numpy as np
import matplotlib.pyplot as plt

# dane
file = "marks_all.xlsx"
data = []

# odczytanie danych
dataframe_raw = openpyxl.load_workbook(f'Data/{file}')
dataframe = dataframe_raw.active

n = dataframe.max_row

for row in range(1, n):
    for col in dataframe.iter_cols(0, dataframe.max_column):
        data.append(col[row].value)

print(data)

# histogram słupkowy

plt.hist(data, bins=16, color='skyblue', edgecolor='black')
plt.title("Histogram ocen")
plt.xlabel("Oceny")
plt.ylabel("Ilość zdobytych ocen")

plt.show()
